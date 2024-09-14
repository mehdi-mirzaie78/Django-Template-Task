from typing import List, Dict
from django.core.exceptions import BadRequest
from openpyxl import load_workbook
from .models import Passenger

field_mapper = {'PassengerId': 'id', 'Survived': 'survived', 'Pclass': 'pclass',
                'Name': 'name', 'Sex': 'sex', 'Age': 'age', 'SibSp': 'sibsp',
                'Parch': 'parch', 'Ticket': 'ticket', 'Fare': 'fare',
                'Cabin': 'cabin', 'Embarked': 'embarked'}


def extract_null_fields(fields: Dict) -> List:
    """
    Extracts fields that doesn't match table's column
    :param fields:
    :return:
    """
    return [key for key, val in fields.items() if val is None]


def extract_and_check_fields(sheet) -> List:
    """
    Iterates over the first row and extracts the fields
    and checks if the field exists on the database if one or more fields doesn't exist
    it will raise an error
    :param sheet:
    :return:
    """
    first_row = None

    for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        first_row = value

    fields = {field: field_mapper.get(field, None) for field in first_row if first_row is not None}

    if not all(fields.values()):
        null_fields = extract_null_fields(fields)
        null_fields_string = ', '.join(null_fields)
        if len(null_fields) <= 1:
            raise BadRequest(f"Wrong dataset: There is no such column as {null_fields_string}.")
        raise BadRequest(f"Wrong dataset: There are no such columns as {null_fields_string}.")
    return list(fields.values())


def insert_data_from_dataset(dataset_file):
    workbook = load_workbook(filename=dataset_file)
    sheet = workbook.active
    passengers = []

    fields = extract_and_check_fields(sheet)

    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        kwargs = dict(zip(fields, row_values))
        kwargs['age'] = None if kwargs['age'] == '' else float(kwargs['age'])
        passengers.append(Passenger(**kwargs))

    return Passenger.objects.bulk_create(passengers)
